import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import openpyxl
excelfile_location=r"E:\test.xlsx"
xlfile=openpyxl.load_workbook(excelfile_location)
print(type(xlfile))
currentsheet1=xlfile.active
print(type(currentsheet1))
cell2_1=currentsheet1._get_cell(2,1)
un1=cell2_1.value
print(un1)
cell2_2=currentsheet1._get_cell(2,2)
ps1=cell2_2.value
print(ps1)
total_rows=currentsheet1.max_row
print("total rows--->",total_rows)
totalColoumns=currentsheet1.max_column
print("total coloumns--->",totalColoumns)
driver=webdriver.Chrome()
driver.set_page_load_timeout(20)
driver.get("https://www.facebook.com/")
wait=WebDriverWait(driver,10)
username=wait.until(ec.presence_of_element_located((By.ID,'email')))
username.send_keys(un1)
passWord=wait.until(ec.presence_of_element_located((By.ID,'pass')))
passWord.send_keys(ps1)
login_btn=wait.until(ec.element_to_be_clickable((By.NAME,"login")))
login_btn.click()
time.sleep(20)
driver.quit()