import openpyxl
path="E:\writedata.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
print("rows--->",rows)
cols=sheet.max_column
print("cols-->",cols)
for r in range(1,11):
    for c in range(1,3):
        sheet.cell(row=r,column=c).value='python'
workbook.save(path)
for r in range(11,21):
    for c in range(1,3):
        sheet.cell(row=r,column=c).value='class'

value1=sheet.cell(row=2,column=2).value="hello"
workbook.save(path)

rows=sheet.max_row
print("after write rows--->",rows)
cols=sheet.max_column
print("cols-->",cols)
for row in range(1,rows+1):
    for col in range(1,cols+1):
        print(sheet.cell(row=row,column=cols).value,end="    ")
    print()