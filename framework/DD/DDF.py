# read the data using for loop
from selenium import  webdriver
import openpyxl
filelocation=r"E:\test.xlsx"
workbook=openpyxl.load_workbook(filelocation)
sheet1=workbook.get_sheet_by_name('register')
rows=sheet1.max_row
print("max rows--->",rows)
colums=sheet1.max_column
print("max column--->",colums)
for r in range(1,rows+1):
    for c in range(1,colums+1):
        print(sheet1.cell(row=r,column=c).value,end="       ")
    print()
