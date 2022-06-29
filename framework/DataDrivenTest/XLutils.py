import openpyxl

def CountRow(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_row
def CountCloumn(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.max_column
def ReadData(file,sheetname,rownum,columnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum,column=columnum).value
def WriteData(file,sheetname,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetname)
    wr=sheet.cell(row=rownum,column=columnnum,).value = data
    workbook.save(file)
    return wr
